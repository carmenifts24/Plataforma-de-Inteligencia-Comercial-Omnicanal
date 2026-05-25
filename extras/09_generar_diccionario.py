"""
Genera el diccionario de datos en Excel para RetailIQ 360.

Salida:
    extras/diccionario_datos_RetailIQ360.xlsx

Version actualizada segun notebooks/04_ETL.ipynb y el archivo vigente de
Power BI. El script toma la estructura real de los CSV del proyecto para no
depender de conteos escritos a mano.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. RUTAS
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parents[1]
EXTRAS_DIR = BASE_DIR / "extras"
DATOS_DIR = BASE_DIR / "datos"
POWER_BI_DIR = BASE_DIR / "power_bi"
OUT_PATH = EXTRAS_DIR / "diccionario_datos_RetailIQ360.xlsx"


# ---------------------------------------------------------------------------
# 2. ESTILO
# ---------------------------------------------------------------------------

C_HEADER = "1F2D3D"
C_SUBHEADER = "2C3E50"
C_DIM = "1A5276"
C_FACT = "7B241C"
C_CACE = "7D6608"
C_REL = "145A32"
C_VERIFY = "4A235A"
C_INFO = "34495E"
C_LIGHT_DIM = "D6EAF8"
C_LIGHT_FACT = "FADBD8"
C_LIGHT_CACE = "FEF9E7"
C_LIGHT_INFO = "EAECEE"
C_WHITE = "FFFFFF"
C_PK = "F9E79F"
C_FK = "AED6F1"
C_CALC = "A9DFBF"
C_WARN = "FDEBD0"

BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def font(color: str = "000000", bold: bool = False, italic: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", color=color, bold=bold, italic=italic, size=size)


def align(h: str = "left", v: str = "center", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def style_cell(cell, bg: str = C_WHITE, bold: bool = False, italic: bool = False, h: str = "left"):
    cell.fill = fill(bg)
    cell.font = font(bold=bold, italic=italic)
    cell.alignment = align(h=h)
    cell.border = BORDER
    return cell


def title(ws, cell_range: str, text: str, color: str = C_HEADER):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = fill(color)
    cell.font = font(C_WHITE, bold=True, size=14)
    cell.alignment = align(h="center")
    cell.border = BORDER


def subtitle(ws, cell_range: str, text: str, color: str = C_SUBHEADER):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = fill(color)
    cell.font = font(C_WHITE, italic=True, size=10)
    cell.alignment = align(h="center")
    cell.border = BORDER


def set_widths(ws, widths: dict[str, int]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ---------------------------------------------------------------------------
# 3. METADATOS DEL MODELO ACTUAL
# ---------------------------------------------------------------------------

TABLES = [
    {
        "nombre": "dim_geografia_ar",
        "categoria": "Dimension",
        "subcategoria": "AR sintetico",
        "archivo": "001_dim_geografia_ar.csv",
        "carpeta": "datos/03_sinteticos",
        "descripcion": "Jerarquia geografica argentina: region, provincia, ciudad, zona y peso relativo de facturacion.",
        "uso_powerbi": "Dimension geografica para mapas, regiones, provincias y relacion con clientes/sucursales.",
    },
    {
        "nombre": "dim_sucursales_ar",
        "categoria": "Dimension",
        "subcategoria": "AR sintetico",
        "archivo": "002_dim_sucursales_ar.csv",
        "carpeta": "datos/03_sinteticos",
        "descripcion": "Red de sucursales fisicas con ubicacion, superficie, empleados y estado de actividad.",
        "uso_powerbi": "Filtra ventas fisicas. La FK SucursalID en ventas queda en blanco para canales digitales.",
    },
    {
        "nombre": "dim_canal_ar",
        "categoria": "Dimension",
        "subcategoria": "AR sintetico",
        "archivo": "003_dim_canal_ar.csv",
        "carpeta": "datos/03_sinteticos",
        "descripcion": "Canales de venta: tienda fisica, web, app y marketplace, calibrados con benchmarks CACE.",
        "uso_powerbi": "Segmentacion omnicanal y comparacion de ventas por canal.",
    },
    {
        "nombre": "dim_clientes_ar",
        "categoria": "Dimension",
        "subcategoria": "Procesado ETL",
        "archivo": "dim_clientes_ar.csv",
        "carpeta": "datos/04_procesados",
        "descripcion": "Clientes sinteticos argentinos enriquecidos con GeografiaID agregado por el ETL.",
        "uso_powerbi": "Segmentacion por cliente, NSE, edad, genero, canal preferido y geografia.",
    },
    {
        "nombre": "dim_inflacion_ipc",
        "categoria": "Dimension",
        "subcategoria": "Procesado ETL",
        "archivo": "dim_inflacion_ipc.csv",
        "carpeta": "datos/04_procesados",
        "descripcion": "Serie mensual de IPC INDEC con PeriodoID e indice_ipc_acum incorporados.",
        "uso_powerbi": "Relacion por PeriodoID con ventas y precios para analisis nominal vs real.",
    },
    {
        "nombre": "dim_tiempo",
        "categoria": "Dimension",
        "subcategoria": "Procesado ETL",
        "archivo": "dim_tiempo.csv",
        "carpeta": "datos/04_procesados",
        "descripcion": "Tabla calendario generada por el ETL. Cubre 2016-2024 para FactVentasFinal y FactPreciosComp.",
        "uso_powerbi": "Debe marcarse como Date Table usando la columna fecha.",
    },
    {
        "nombre": "dim_productos",
        "categoria": "Dimension",
        "subcategoria": "Procesado ETL",
        "archivo": "dim_productos.csv",
        "carpeta": "datos/04_procesados",
        "descripcion": "Dimension de productos Olist con categoria original, categoria traducida, categoria CACE, peso, dimensiones y volumen.",
        "uso_powerbi": "Analisis por producto/categoria y relacion entre atributos fisicos y flete.",
    },
    {
        "nombre": "dim_categorias",
        "categoria": "Dimension",
        "subcategoria": "Procesado ETL",
        "archivo": "dim_categorias.csv",
        "carpeta": "datos/04_procesados",
        "descripcion": "Tabla puente de categorias CACE generada para evitar relaciones N:N entre productos y precios de competencia.",
        "uso_powerbi": "Relaciona dim_productos y fact_precios_comp por categoria_es/categoria.",
    },
    {
        "nombre": "fact_ventas_final",
        "categoria": "Tabla de hechos",
        "subcategoria": "Procesado ETL",
        "archivo": "fact_ventas_final.csv",
        "carpeta": "datos/04_procesados",
        "descripcion": "Tabla principal de ventas Olist convertidas a ARS, ajustadas por IPC y enriquecidas con contexto argentino.",
        "uso_powerbi": "Fact table principal para KPIs comerciales, ventas nominales/reales, canal, cliente, producto y logistica.",
    },
    {
        "nombre": "fact_precios_comp",
        "categoria": "Tabla de hechos",
        "subcategoria": "Procesado ETL",
        "archivo": "fact_precios_comp.csv",
        "carpeta": "datos/04_procesados",
        "descripcion": "Precios propios y de competencia con PeriodoID agregado por el ETL.",
        "uso_powerbi": "Modulo de pricing: price index, descuento, comparacion con competencia y evolucion temporal.",
    },
]


CACE_TABLES = [
    ("cace_00_README", "cace_00_README.csv", "Inventario de archivos CACE y usos principales."),
    ("cace_01_kpis_macro", "cace_01_kpis_macro.csv", "KPIs macro del e-commerce argentino: facturacion, ordenes, ticket, conversion, inflacion y crecimiento."),
    ("cace_02_categorias_rubros", "cace_02_categorias_rubros.csv", "Rubros/categorias del e-commerce argentino con facturacion, participacion y rankings."),
    ("cace_03a_conversion_por_categoria", "cace_03a_conversion_por_categoria.csv", "Tasas de conversion por categoria."),
    ("cace_03b_ranking_demanda", "cace_03b_ranking_demanda.csv", "Ranking de demanda por categoria."),
    ("cace_04a_medios_pago_oferta", "cace_04a_medios_pago_oferta.csv", "Oferta de medios de pago y evolucion porcentual."),
    ("cace_04b_financiamiento_cuotas", "cace_04b_financiamiento_cuotas.csv", "Distribucion de ventas por plazo de cuotas."),
    ("cace_05a_logistica_tipo_entrega", "cace_05a_logistica_tipo_entrega.csv", "Distribucion de tipos de entrega."),
    ("cace_05b_plazos_entrega", "cace_05b_plazos_entrega.csv", "Plazos de entrega y diferencias AMBA/interior."),
    ("cace_06a_distribucion_regional", "cace_06a_distribucion_regional.csv", "Distribucion regional de facturacion."),
    ("cace_06b_perfil_comprador", "cace_06b_perfil_comprador.csv", "Perfil del comprador por variable y segmento."),
    ("cace_07_canal_online_por_categoria", "cace_07_canal_online_por_categoria.csv", "Peso del canal online por categoria."),
]

for nombre, archivo, descripcion in CACE_TABLES:
    TABLES.append(
        {
            "nombre": nombre,
            "categoria": "Benchmark CACE",
            "subcategoria": "Referencia",
            "archivo": archivo,
            "carpeta": "datos/02_cace_benchmarks",
            "descripcion": descripcion,
            "uso_powerbi": "Tabla de referencia sin relaciones formales; se usa para comparaciones y benchmarks.",
        }
    )


DEPRECATED_OR_SOURCE_TABLES = [
    {
        "nombre": "004_dim_clientes_ar",
        "archivo": "004_dim_clientes_ar.csv",
        "carpeta": "datos/03_sinteticos",
        "estado": "Origen reemplazado",
        "nota": "Usar en Power BI la version procesada datos/04_procesados/dim_clientes_ar.csv, que incluye GeografiaID.",
    },
    {
        "nombre": "005_fact_ventas_base_ar",
        "archivo": "005_fact_ventas_base_ar.csv",
        "carpeta": "datos/03_sinteticos",
        "estado": "Deprecada para el modelo final",
        "nota": "Su contexto argentino fue integrado en fact_ventas_final. No cargar como tabla de hechos principal.",
    },
    {
        "nombre": "006_fact_precios_comp",
        "archivo": "006_fact_precios_comp.csv",
        "carpeta": "datos/03_sinteticos",
        "estado": "Origen reemplazado",
        "nota": "Usar datos/04_procesados/fact_precios_comp.csv, que incluye PeriodoID para relaciones temporales.",
    },
]


RELACIONES = [
    ("fact_ventas_final", "fecha", "dim_tiempo", "fecha", "N:1", "Activa", "Marcar dim_tiempo como Date Table."),
    ("fact_ventas_final", "CanalID", "dim_canal_ar", "CanalID", "N:1", "Activa", "Relacion de canal principal."),
    ("fact_ventas_final", "ClienteID", "dim_clientes_ar", "ClienteID", "N:1", "Activa", "Relacion de cliente."),
    ("fact_ventas_final", "SucursalID", "dim_sucursales_ar", "SucursalID", "N:1", "Activa", "SucursalID queda en blanco para canales digitales."),
    ("fact_ventas_final", "PeriodoID", "dim_inflacion_ipc", "PeriodoID", "N:1", "Activa", "No crear anio_mes calculado; usar PeriodoID."),
    ("fact_ventas_final", "product_id", "dim_productos", "product_id", "N:1", "Activa", "Relacion por producto individual, no por categoria."),
    ("dim_clientes_ar", "GeografiaID", "dim_geografia_ar", "GeografiaID", "N:1", "Activa", "Habilita analisis geografico de clientes."),
    ("dim_sucursales_ar", "GeografiaID", "dim_geografia_ar", "GeografiaID", "N:1", "Activa", "Habilita analisis geografico de sucursales."),
    ("fact_precios_comp", "CanalID", "dim_canal_ar", "CanalID", "N:1", "Activa", "Canal asociado al precio relevado."),
    ("fact_precios_comp", "PeriodoID", "dim_inflacion_ipc", "PeriodoID", "N:1", "Activa", "Relacion mensual por PeriodoID."),
    ("fact_precios_comp", "fecha_relevamiento", "dim_tiempo", "fecha", "N:1", "Manual", "Crear manualmente porque las columnas tienen nombres distintos."),
    ("fact_precios_comp", "categoria", "dim_categorias", "categoria_es", "N:1", "Manual", "Crear manualmente porque las columnas tienen nombres distintos."),
    ("dim_productos", "categoria_es", "dim_categorias", "categoria_es", "N:1", "Activa", "DimCategorias evita relacion N:N con FactPreciosComp."),
]


VERIFICACIONES = [
    ("dim_tiempo", "Formato", "Marcar como tabla de fechas usando dim_tiempo[fecha]."),
    ("dim_tiempo", "Integridad", "Debe cubrir 2016-01-01 a 2024-12-31 para ventas y precios de competencia."),
    ("dim_clientes_ar", "Integridad", "ClienteID debe ser unico y todos los ClienteID de fact_ventas_final deben existir."),
    ("dim_clientes_ar", "Integridad", "GeografiaID no debe tener nulos."),
    ("dim_inflacion_ipc", "Integridad", "PeriodoID debe ser unico y cubrir los PeriodoID usados por fact_ventas_final y fact_precios_comp."),
    ("dim_productos", "Integridad", "product_id debe ser unico y cubrir todos los product_id de fact_ventas_final."),
    ("dim_productos", "Negocio", "categoria_es debe tener categorias CACE consistentes para vincular con fact_precios_comp."),
    ("dim_categorias", "Integridad", "categoria_es debe ser unico; es la tabla puente entre productos y precios."),
    ("fact_ventas_final", "Integridad", "Debe tener 110.197 filas y 28 columnas."),
    ("fact_ventas_final", "Negocio", "precio_venta_ars y precio_venta_ars_real deben ser positivos."),
    ("fact_ventas_final", "Negocio", "SucursalID puede estar en blanco para canales digitales; no corregirlo como error."),
    ("fact_precios_comp", "Integridad", "Debe tener PeriodoID y 360 filas procesadas."),
    ("fact_precios_comp", "Negocio", "price_index menor que 1 indica precio propio mas barato que competencia."),
    ("Benchmarks CACE", "Modelo", "No crear relaciones formales automaticas con tablas CACE salvo decision explicita de visualizacion."),
    ("Power BI", "Modelo", "Eliminar relaciones automaticas incorrectas, especialmente por category_en o columnas de texto similares."),
]


CLAVES = {
    "dim_geografia_ar": {"GeografiaID": ("PK", None, None)},
    "dim_sucursales_ar": {
        "SucursalID": ("PK", None, None),
        "GeografiaID": ("FK", "dim_geografia_ar", "GeografiaID"),
    },
    "dim_canal_ar": {"CanalID": ("PK", None, None)},
    "dim_clientes_ar": {
        "ClienteID": ("PK", None, None),
        "GeografiaID": ("FK", "dim_geografia_ar", "GeografiaID"),
    },
    "dim_inflacion_ipc": {"PeriodoID": ("PK", None, None), "indice_ipc_acum": ("CALC", None, None)},
    "dim_tiempo": {"fecha": ("PK", None, None), "TiempoID": ("PK", None, None), "PeriodoID": ("CALC", None, None)},
    "dim_productos": {"product_id": ("PK", None, None), "categoria_es": ("FK", "dim_categorias", "categoria_es"), "volumen_cm3": ("CALC", None, None)},
    "dim_categorias": {"CategoriaID": ("PK", None, None), "categoria_es": ("PK", None, None)},
    "fact_ventas_final": {
        "VentaID": ("PK", None, None),
        "ClienteID": ("FK", "dim_clientes_ar", "ClienteID"),
        "CanalID": ("FK", "dim_canal_ar", "CanalID"),
        "SucursalID": ("FK", "dim_sucursales_ar", "SucursalID"),
        "PeriodoID": ("FK", "dim_inflacion_ipc", "PeriodoID"),
        "fecha": ("FK", "dim_tiempo", "fecha"),
        "product_id": ("FK", "dim_productos", "product_id"),
        "tipo_cambio": ("CALC", None, None),
        "precio_venta_ars": ("CALC", None, None),
        "flete_ars": ("CALC", None, None),
        "indice_ipc_acum": ("CALC", None, None),
        "precio_venta_ars_real": ("CALC", None, None),
    },
    "fact_precios_comp": {
        "PrecioID": ("PK", None, None),
        "CanalID": ("FK", "dim_canal_ar", "CanalID"),
        "PeriodoID": ("FK", "dim_inflacion_ipc", "PeriodoID"),
        "fecha_relevamiento": ("FK", "dim_tiempo", "fecha"),
        "categoria": ("FK", "dim_categorias", "categoria_es"),
        "price_index": ("CALC", None, None),
    },
}


DESCRIPCIONES_COLUMNAS = {
    "VentaID": "Clave unica de venta generada por el ETL.",
    "ClienteID": "Clave del cliente sintetico argentino.",
    "CanalID": "Clave del canal de venta.",
    "SucursalID": "Clave de sucursal; puede estar en blanco en ventas digitales.",
    "PeriodoID": "Clave mensual numerica: anio * 100 + mes.",
    "fecha": "Fecha de la transaccion o del periodo.",
    "anio": "Anio del registro.",
    "mes": "Mes numerico del registro.",
    "trimestre": "Trimestre del registro.",
    "category_en": "Categoria de producto traducida al ingles.",
    "categoria_es": "Categoria simplificada en espanol alineada a CACE.",
    "product_id": "Identificador original del producto Olist.",
    "product_category_name": "Categoria original Olist en portugues.",
    "product_weight_g": "Peso del producto en gramos.",
    "product_length_cm": "Largo del producto en centimetros.",
    "product_height_cm": "Alto del producto en centimetros.",
    "product_width_cm": "Ancho del producto en centimetros.",
    "volumen_cm3": "Volumen calculado: largo * alto * ancho.",
    "medio_pago": "Medio de pago asignado con distribucion calibrada por CACE.",
    "nro_cuotas": "Cantidad de cuotas asignada con distribucion calibrada por CACE.",
    "tipo_entrega": "Tipo de entrega asignado con benchmarks CACE.",
    "plazo_entrega": "Plazo de entrega asignado con benchmarks CACE.",
    "cantidad": "Unidades vendidas.",
    "precio_venta_brl": "Precio original de Olist en reales brasilenos.",
    "flete_brl": "Flete original de Olist en reales brasilenos.",
    "ars_por_usd": "Tipo de cambio ARS/USD usado como moneda puente.",
    "tipo_cambio": "Factor BRL a ARS calculado como ars_por_usd / 3.3.",
    "precio_venta_ars": "Precio nominal en pesos argentinos.",
    "flete_ars": "Flete nominal en pesos argentinos.",
    "ipc_nivel_general": "Variacion mensual del IPC general.",
    "indice_ipc_acum": "Indice acumulado usado para deflactar precios.",
    "precio_venta_ars_real": "Precio en pesos argentinos reales, deflactado por IPC.",
    "tiene_ipc": "Indica si la fila tiene cobertura de IPC para el periodo.",
    "order_id": "Identificador original de orden Olist.",
    "seller_id": "Identificador original del vendedor Olist.",
    "PrecioID": "Clave del registro de precio de competencia.",
    "fecha_relevamiento": "Fecha del relevamiento de precios.",
    "precio_lista_ars": "Precio de lista propio en ARS.",
    "precio_web_ars": "Precio online propio en ARS.",
    "precio_competencia": "Precio promedio de competencia en ARS.",
    "descuento_pct": "Porcentaje de descuento aplicado.",
    "price_index": "Indice precio propio / precio competencia.",
    "GeografiaID": "Clave geografica para relacionar con dim_geografia_ar.",
    "TiempoID": "Clave de fecha en formato YYYYMMDD.",
}


# ---------------------------------------------------------------------------
# 4. UTILIDADES DE DATOS
# ---------------------------------------------------------------------------

def csv_path(table: dict) -> Path:
    return BASE_DIR / table["carpeta"] / table["archivo"]


def read_header(path: Path) -> pd.DataFrame:
    try:
        return pd.read_csv(path, encoding="utf-8-sig", nrows=200)
    except pd.errors.ParserError:
        return pd.read_csv(path, encoding="utf-8-sig", nrows=200, engine="python", on_bad_lines="skip")


def count_rows(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", errors="replace") as fh:
        return max(sum(1 for _ in fh) - 1, 0)


def infer_type(series: pd.Series) -> str:
    if pd.api.types.is_bool_dtype(series):
        return "Booleano"
    if pd.api.types.is_integer_dtype(series):
        return "Entero"
    if pd.api.types.is_float_dtype(series):
        return "Decimal"
    if "fecha" in series.name.lower() or series.name.lower() == "date":
        return "Fecha"
    return "Texto"


def enrich_tables() -> list[dict]:
    enriched = []
    for table in TABLES:
        path = csv_path(table)
        table = table.copy()
        if path.exists():
            sample = read_header(path)
            table["filas"] = count_rows(path)
            table["columnas_count"] = len(sample.columns)
            table["existe"] = "Si"
            cols = []
            for col in sample.columns:
                clave, rel_table, rel_col = CLAVES.get(table["nombre"], {}).get(col, ("", None, None))
                cols.append(
                    {
                        "nombre": col,
                        "tipo": infer_type(sample[col]),
                        "clave": clave,
                        "tabla_rel": rel_table,
                        "col_rel": rel_col,
                        "descripcion": DESCRIPCIONES_COLUMNAS.get(col, f"Columna {col} de {table['nombre']}."),
                    }
                )
            table["columnas"] = cols
        else:
            table["filas"] = 0
            table["columnas_count"] = 0
            table["existe"] = "No"
            table["columnas"] = []
        enriched.append(table)
    return enriched


def category_colors(categoria: str) -> tuple[str, str]:
    if categoria == "Dimension":
        return C_DIM, C_LIGHT_DIM
    if categoria == "Tabla de hechos":
        return C_FACT, C_LIGHT_FACT
    if categoria == "Benchmark CACE":
        return C_CACE, C_LIGHT_CACE
    return C_INFO, C_LIGHT_INFO


# ---------------------------------------------------------------------------
# 5. HOJAS
# ---------------------------------------------------------------------------

def build_indice(wb, tables: list[dict]):
    ws = wb.create_sheet("Indice de tablas")
    ws.sheet_view.showGridLines = False
    title(ws, "A1:J1", "RetailIQ 360 - Diccionario de Datos")
    subtitle(
        ws,
        "A2:J2",
        "Modelo actualizado desde notebooks/04_ETL.ipynb: FactVentasFinal, FactPreciosComp, DimProductos, DimCategorias y Date Table.",
    )

    headers = ["#", "Tabla", "Categoria", "Subcategoria", "Archivo", "Carpeta", "Existe", "Filas", "Columnas", "Descripcion"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_cell(c, C_SUBHEADER, bold=True, h="center")
        c.font = font(C_WHITE, bold=True)

    for idx, table in enumerate(tables, 1):
        row = idx + 4
        _, bg = category_colors(table["categoria"])
        values = [
            idx,
            table["nombre"],
            table["categoria"],
            table["subcategoria"],
            table["archivo"],
            table["carpeta"],
            table["existe"],
            table["filas"],
            table["columnas_count"],
            table["descripcion"],
        ]
        for col, value in enumerate(values, 1):
            style_cell(ws.cell(row=row, column=col, value=value), bg, bold=(col == 2), h="center" if col in (1, 3, 7, 8, 9) else "left")
        ws.row_dimensions[row].height = 34

    set_widths(ws, {"A": 5, "B": 30, "C": 18, "D": 18, "E": 38, "F": 28, "G": 8, "H": 12, "I": 10, "J": 70})
    ws.freeze_panes = "A5"


def build_diccionario(wb, tables: list[dict]):
    ws = wb.create_sheet("Diccionario completo")
    ws.sheet_view.showGridLines = False
    title(ws, "A1:K1", "RetailIQ 360 - Diccionario Completo")
    subtitle(ws, "A2:K2", "PK = clave primaria | FK = clave foranea | CALC = campo calculado o derivado por ETL")

    headers = ["Tabla", "Categoria", "#", "Columna", "Tipo", "Clave", "Tabla relacionada", "Columna relacionada", "Descripcion", "Archivo", "Uso Power BI"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_cell(c, C_SUBHEADER, bold=True, h="center")
        c.font = font(C_WHITE, bold=True)

    row = 4
    key_bg = {"PK": C_PK, "FK": C_FK, "CALC": C_CALC, "": C_WHITE}

    for table in tables:
        header_color, bg = category_colors(table["categoria"])
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        c = ws.cell(row=row, column=1, value=f"{table['nombre']} - {table['descripcion']}")
        style_cell(c, header_color, bold=True)
        c.font = font(C_WHITE, bold=True)

        for idx, col_meta in enumerate(table["columnas"], 1):
            row += 1
            values = [
                table["nombre"],
                table["categoria"],
                idx,
                col_meta["nombre"],
                col_meta["tipo"],
                col_meta["clave"] or "-",
                col_meta["tabla_rel"] or "-",
                col_meta["col_rel"] or "-",
                col_meta["descripcion"],
                table["archivo"],
                table["uso_powerbi"],
            ]
            for col_idx, value in enumerate(values, 1):
                cell_bg = key_bg.get(col_meta["clave"], bg) if col_idx == 6 else bg
                style_cell(
                    ws.cell(row=row, column=col_idx, value=value),
                    cell_bg,
                    bold=(col_idx == 4 and col_meta["clave"] == "PK") or col_idx == 1,
                    h="center" if col_idx in (3, 5, 6) else "left",
                )
            ws.row_dimensions[row].height = 30

    set_widths(
        ws,
        {
            "A": 28,
            "B": 18,
            "C": 5,
            "D": 30,
            "E": 12,
            "F": 9,
            "G": 26,
            "H": 22,
            "I": 62,
            "J": 34,
            "K": 60,
        },
    )
    ws.freeze_panes = "A5"


def build_relaciones(wb):
    ws = wb.create_sheet("Relaciones modelo")
    ws.sheet_view.showGridLines = False
    title(ws, "A1:H1", "RetailIQ 360 - Relaciones del Modelo", C_REL)
    subtitle(ws, "A2:H2", "Esquema galaxia actualizado para Power BI. Las tablas CACE quedan como benchmarks sin relaciones formales.", C_REL)

    headers = ["#", "Tabla origen", "Columna origen", "Tabla destino", "Columna destino", "Cardinalidad", "Tipo", "Notas"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_cell(c, C_SUBHEADER, bold=True, h="center")
        c.font = font(C_WHITE, bold=True)

    for idx, rel in enumerate(RELACIONES, 1):
        row = idx + 4
        bg = C_LIGHT_FACT if rel[0].startswith("fact_") else C_LIGHT_DIM
        if rel[5] == "Manual":
            bg = C_WARN
        values = [idx, *rel]
        for col, value in enumerate(values, 1):
            style_cell(ws.cell(row=row, column=col, value=value), bg, bold=(col in (2, 3)), h="center" if col in (1, 6, 7) else "left")
        ws.row_dimensions[row].height = 30

    note_row = len(RELACIONES) + 7
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    note = ws.cell(
        row=note_row,
        column=1,
        value="Importante: verificar que Power BI no cree relaciones automaticas incorrectas por columnas de texto como category_en. La relacion de productos debe usar product_id.",
    )
    style_cell(note, C_WARN, italic=True)
    ws.row_dimensions[note_row].height = 38

    set_widths(ws, {"A": 5, "B": 28, "C": 24, "D": 28, "E": 24, "F": 14, "G": 12, "H": 70})
    ws.freeze_panes = "A5"


def build_verificaciones(wb):
    ws = wb.create_sheet("Verificaciones")
    ws.sheet_view.showGridLines = False
    title(ws, "A1:E1", "RetailIQ 360 - Checklist Power BI", C_VERIFY)

    headers = ["Estado", "Tabla", "Tipo", "Verificacion", "Observaciones"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        style_cell(c, C_SUBHEADER, bold=True, h="center")
        c.font = font(C_WHITE, bold=True)

    tipo_bg = {"Integridad": "D5E8D4", "Negocio": "DAE8FC", "Formato": "FFF2CC", "Modelo": C_WARN}
    for idx, (tabla, tipo, texto) in enumerate(VERIFICACIONES, 1):
        row = idx + 3
        bg = tipo_bg.get(tipo, C_WHITE)
        values = ["☐", tabla, tipo, texto, ""]
        for col, value in enumerate(values, 1):
            style_cell(ws.cell(row=row, column=col, value=value), bg if col < 5 else C_WHITE, h="center" if col in (1, 3) else "left")
        ws.row_dimensions[row].height = 25

    set_widths(ws, {"A": 8, "B": 28, "C": 14, "D": 85, "E": 45})
    ws.freeze_panes = "A4"


def build_orden_carga(wb, tables: list[dict]):
    ws = wb.create_sheet("Orden carga Power BI")
    ws.sheet_view.showGridLines = False
    title(ws, "A1:F1", "RetailIQ 360 - Orden de Carga en Power BI")
    subtitle(ws, "A2:F2", "Cargar dimensiones, luego hechos procesados, y al final benchmarks CACE sin relaciones formales.")

    order = [
        "dim_geografia_ar",
        "dim_canal_ar",
        "dim_sucursales_ar",
        "dim_clientes_ar",
        "dim_inflacion_ipc",
        "dim_tiempo",
        "dim_categorias",
        "dim_productos",
        "fact_ventas_final",
        "fact_precios_comp",
    ]
    table_map = {t["nombre"]: t for t in tables}
    ordered = [table_map[name] for name in order if name in table_map]
    ordered += [t for t in tables if t["categoria"] == "Benchmark CACE"]

    headers = ["Paso", "Tabla", "Archivo", "Carpeta", "Tipo", "Notas"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_cell(c, C_SUBHEADER, bold=True, h="center")
        c.font = font(C_WHITE, bold=True)

    notes = {
        "dim_tiempo": "Marcar como Date Table con la columna fecha.",
        "dim_clientes_ar": "Usar la version procesada porque incluye GeografiaID.",
        "dim_inflacion_ipc": "Usar PeriodoID para relacionar; no crear anio_mes calculado.",
        "dim_categorias": "Cargar antes de dim_productos/fact_precios_comp para evitar N:N.",
        "dim_productos": "Relacionar con fact_ventas_final por product_id y con dim_categorias por categoria_es.",
        "fact_ventas_final": "Tabla de hechos principal. No cargar fact_ventas_base_ar como sustituto.",
        "fact_precios_comp": "Usar version procesada con PeriodoID.",
    }

    for idx, table in enumerate(ordered, 1):
        row = idx + 4
        _, bg = category_colors(table["categoria"])
        note = notes.get(table["nombre"], "Benchmark de referencia sin relaciones formales." if table["categoria"] == "Benchmark CACE" else "")
        values = [idx, table["nombre"], table["archivo"], table["carpeta"], table["categoria"], note]
        for col, value in enumerate(values, 1):
            style_cell(ws.cell(row=row, column=col, value=value), bg, bold=(col in (1, 2)), h="center" if col in (1, 5) else "left")
        ws.row_dimensions[row].height = 28

    set_widths(ws, {"A": 7, "B": 30, "C": 38, "D": 28, "E": 18, "F": 75})
    ws.freeze_panes = "A5"


def build_powerbi(wb):
    ws = wb.create_sheet("Power BI")
    ws.sheet_view.showGridLines = False
    title(ws, "A1:F1", "RetailIQ 360 - Archivos Power BI", C_INFO)
    subtitle(ws, "A2:F2", "Informacion tomada de la carpeta power_bi.")

    headers = ["#", "Archivo", "Tamanio MB", "Modificado", "Estado", "Notas"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_cell(c, C_SUBHEADER, bold=True, h="center")
        c.font = font(C_WHITE, bold=True)

    pbix_files = sorted(POWER_BI_DIR.glob("*.pbix"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not pbix_files:
        row = 5
        ws.merge_cells("A5:F5")
        style_cell(ws["A5"], C_WARN)
        ws["A5"].value = "No se encontraron archivos .pbix en power_bi."
    else:
        latest = pbix_files[0]
        for idx, path in enumerate(pbix_files, 1):
            stat = path.stat()
            row = idx + 4
            values = [
                idx,
                path.name,
                round(stat.st_size / (1024 * 1024), 2),
                datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "Actual" if path == latest else "Version anterior",
                "Archivo vigente para continuar modelado y dashboard." if path == latest else "Mantener solo como respaldo si corresponde.",
            ]
            for col, value in enumerate(values, 1):
                style_cell(ws.cell(row=row, column=col, value=value), C_LIGHT_INFO, bold=(col == 2), h="center" if col in (1, 3, 5) else "left")

    note_row = max(6, len(pbix_files) + 7)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    note = ws.cell(
        row=note_row,
        column=1,
        value="Checklist de modelo: usar los CSV procesados de datos/04_procesados, marcar dim_tiempo como Date Table y revisar relaciones manuales de FactPreciosComp.",
    )
    style_cell(note, C_WARN, italic=True)
    ws.row_dimensions[note_row].height = 36

    set_widths(ws, {"A": 5, "B": 65, "C": 12, "D": 20, "E": 18, "F": 75})


def build_fuentes_deprecadas(wb):
    ws = wb.create_sheet("Fuentes no cargar")
    ws.sheet_view.showGridLines = False
    title(ws, "A1:F1", "RetailIQ 360 - Fuentes de Origen o Tablas Deprecadas", C_INFO)
    subtitle(ws, "A2:F2", "Archivos utiles para trazabilidad, pero no recomendados como tablas finales del modelo Power BI.")

    headers = ["#", "Tabla/archivo", "Archivo", "Carpeta", "Estado", "Nota"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        style_cell(c, C_SUBHEADER, bold=True, h="center")
        c.font = font(C_WHITE, bold=True)

    for idx, item in enumerate(DEPRECATED_OR_SOURCE_TABLES, 1):
        path = BASE_DIR / item["carpeta"] / item["archivo"]
        row = idx + 4
        values = [idx, item["nombre"], item["archivo"], item["carpeta"], item["estado"], item["nota"]]
        for col, value in enumerate(values, 1):
            style_cell(ws.cell(row=row, column=col, value=value), C_WARN, bold=(col in (1, 2)), h="center" if col in (1, 5) else "left")
        if not path.exists():
            ws.cell(row=row, column=6).value = f"{item['nota']} Archivo no encontrado actualmente."
        ws.row_dimensions[row].height = 34

    set_widths(ws, {"A": 5, "B": 30, "C": 38, "D": 28, "E": 22, "F": 85})


def autosize_filter(wb):
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= 4 and max_col >= 1:
            ws.auto_filter.ref = f"A4:{get_column_letter(max_col)}{max_row}"


# ---------------------------------------------------------------------------
# 6. MAIN
# ---------------------------------------------------------------------------

def main():
    tables = enrich_tables()

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_indice(wb, tables)
    build_diccionario(wb, tables)
    build_relaciones(wb)
    build_verificaciones(wb)
    build_orden_carga(wb, tables)
    build_powerbi(wb)
    build_fuentes_deprecadas(wb)
    autosize_filter(wb)

    wb.save(OUT_PATH)
    print(f"Archivo generado: {OUT_PATH}")


if __name__ == "__main__":
    main()
